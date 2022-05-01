# -*- coding: utf-8 -*-
"""
Create Time: 2022/4/30 19:34
Author: Lison Song
"""
import openpyxl
import re


def time_format(timestr):
    '''
    :param timestr: eg: 1小时34分钟43秒
    :return: 转换为秒数返回
    '''
    result = 0
    hour_list = re.findall(r'(\d+)小时', timestr)
    minute_list = re.findall(r'(\d+)分钟', timestr)
    second_list = re.findall(r'(\d+)秒', timestr)
    if len(hour_list):
        result = result + int(hour_list[0]) * 60 * 60
    if len(minute_list):
        result = result + int(minute_list[0]) * 60
    if len(second_list):
        result = result + int(second_list[0])
    return result


def openpyxl_read_test(xl_filepath):
    # print(xl_filepath)
    inv_file = openpyxl.load_workbook(xl_filepath)
    xl_sheet = inv_file["sheet1"]
    # print(xl_sheet)
    for row in range(7, xl_sheet.max_row + 1):
        student = str(xl_sheet.cell(row, 1).value).strip()
        student_class = str(xl_sheet.cell(row, 2).value).strip()
        parents = str(xl_sheet.cell(row, 3).value).strip()
        watch_time = str(xl_sheet.cell(row, 4).value).strip()
        start_time = str(xl_sheet.cell(row, 5).value).strip()
        end_time = str(xl_sheet.cell(row, 6).value).strip()
        is_comments = str(xl_sheet.cell(row, 7).value).strip()
        watch_time = time_format(watch_time)

        print(student, student_class, parents, watch_time, start_time, end_time, is_comments)

def openpyxl_write_test(xl_filepath,new_xls_file):
    inv_file = openpyxl.load_workbook(xl_filepath)
    xl_sheet = inv_file["sheet1"]
    xl_sheet.cell(6, 8).value = "观看时长(秒数)"
    for row in range(7, xl_sheet.max_row + 1):
        watch_time = str(xl_sheet.cell(row, 4).value).strip()
        watch_time = time_format(watch_time)
        new_cell = xl_sheet.cell(row, 8)
        new_cell.value = watch_time
    inv_file.save(new_xls_file)

if __name__ == '__main__':
    filepath = r"D:\coding\mypython\pythonstudy2022\res\4.6的直播课-直播明细.xlsx"
    new_filepath = r"D:\coding\mypython\pythonstudy2022\res\4.6的直播课-直播明细-rewrite.xlsx"
    openpyxl_read_test(filepath)
    openpyxl_write_test(filepath,new_filepath)
